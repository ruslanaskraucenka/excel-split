import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import io

st.title("Excel Splitter with Formatting Preservation")

uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

MAX_ROWS = 1999

def copy_sheet_format(source_ws, target_ws):
    # Copy column widths
    for col in source_ws.column_dimensions:
        target_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width

    # Copy row heights
    for row in source_ws.row_dimensions:
        target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height

    # Copy merged cells
    for merged_cell in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_cell))

    # Copy styles and values
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            if cell.has_style:
                new_cell.font = cell.font
                new_cell.border = cell.border
                new_cell.fill = cell.fill
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection
                new_cell.alignment = cell.alignment

if uploaded_file:
    in_mem_file = io.BytesIO(uploaded_file.read())
    wb = load_workbook(in_mem_file)
    ws = wb.active

    total_rows = ws.max_row
    num_chunks = (total_rows + MAX_ROWS - 1) // MAX_ROWS

    st.success(f"File uploaded successfully! Splitting into {num_chunks} files...")

    for i in range(num_chunks):
        start_row = i * MAX_ROWS + 1
        end_row = min(start_row + MAX_ROWS - 1, total_rows)

        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = ws.title

        copy_sheet_format(ws, new_ws)

        for row in ws.iter_rows(min_row=start_row, max_row=end_row):
            for cell in row:
                new_cell = new_ws.cell(row=cell.row - start_row + 1, column=cell.col_idx, value=cell.value)
                if cell.has_style:
                    new_cell.font = cell.font
                    new_cell.border = cell.border
                    new_cell.fill = cell.fill
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection
                    new_cell.alignment = cell.alignment

        output = io.BytesIO()
        new_wb.save(output)
        output.seek(0)

        filename = f"{uploaded_file.name.replace('.xlsx', '')}_part_{i+1}.xlsx"
        st.download_button(
            label=f"Download {filename}",
            data=output,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
